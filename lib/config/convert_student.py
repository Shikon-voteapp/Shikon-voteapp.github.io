import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os

# Excelからデータを読み込み、Dartファイルを生成する関数
def excel_to_dart(excel_file):
    # Excelファイルを読み込む
    wb = load_workbook(excel_file)
    sheet = wb.active

    # Excelのデータを読み込む（仮にA列が学年、B列がクラス名、C列が番号、D列がランダム番号であると仮定）
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 1行目はヘッダーと仮定
        grade, class_name, number, student_id = row
        if grade and class_name and number and student_id:  # 空のセルがあった場合にスキップ
            data.append((grade, class_name, number, student_id))

    # デバッグ：読み込んだデータを確認
    print("読み込んだデータ:")
    for row in data:
        print(row)

    # Dartコードを生成
    dart_code = "import '../models/student.dart';\n\n"
    dart_code += "class StudentMapping {\n"
    dart_code += "  static Map<String, Student> getMappings() {\n"
    dart_code += "    return {\n"

    for grade, class_name, number, student_id in data:
        dart_code += f"      '{student_id}': Student(grade: '{grade}', className: '{class_name}', number: {number}),\n"

    dart_code += "    };\n"
    dart_code += "  }\n"
    dart_code += "}\n"

    return dart_code

# GUIを作成してExcelファイルを選択
def open_file_dialog():
    root = tk.Tk()
    root.withdraw()  # メインウィンドウを隠す
    file_path = filedialog.askopenfilename(title="Excelファイルを選択", filetypes=[("Excel Files", "*.xlsx;*.xls")])
    return file_path

# メイン処理
def main():
    # Excelファイルを選択
    excel_file = open_file_dialog()

    if excel_file:
        # Dartコードを生成
        dart_code = excel_to_dart(excel_file)


        # 上書き前に既存のファイルを削除
        dart_file_path = os.path.join('student_mapping.dart')
        if os.path.exists(dart_file_path):
            os.remove(dart_file_path)  # 既存のファイルを削除

        # 新しいDartコードをファイルに保存
        with open(dart_file_path, 'w', encoding='utf-8') as file:
            file.write(dart_code)

        print(f"Dartファイルが生成されました: {dart_file_path}")
    else:
        print("ファイルが選択されませんでした")

# 実行
if __name__ == "__main__":
    main()

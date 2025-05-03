import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def convert_excel_to_dart():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return

    try:
        wb = load_workbook(file_path)
        sheet = wb.active  # 最初のシートを使用
        uuid_list = []

        for row in sheet.iter_rows(min_row=2, max_col=1):  # ヘッダー行をスキップ
            cell = row[0]
            if cell.value is not None:
                try:
                    uuid = int(cell.value)
                    uuid_list.append(uuid)
                except ValueError:
                    pass  # 数値でない場合はスキップ

        if not uuid_list:
            messagebox.showerror("エラー", "UUIDが見つかりませんでした。A列に整数を入力してください。")
            return

        # Dartコード生成
        dart_code = "const List<int> validUuids = [\n"
        dart_code += ",\n".join(f"  {uuid}" for uuid in uuid_list)
        dart_code += "\n];\n"

        save_path = filedialog.asksaveasfilename(defaultextension=".dart", filetypes=[("Dart Files", "*.dart")])
        if save_path:
            with open(save_path, "w", encoding="utf-8") as f:
                f.write(dart_code)
            messagebox.showinfo("成功", "valid_uuids.dart を保存しました。")

    except Exception as e:
        messagebox.showerror("エラー", f"変換に失敗しました:\n{str(e)}")

# GUIセットアップ
root = tk.Tk()
root.title("Excel → Dart UUID 変換ツール")
root.geometry("300x150")

label = tk.Label(root, text="Excelファイルを選んでDart形式に変換")
label.pack(pady=20)

btn = tk.Button(root, text="Excelファイルを選択", command=convert_excel_to_dart)
btn.pack()

root.mainloop()
